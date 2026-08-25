#!/usr/bin/env python3
"""Extract data/union-budget.json from the vendored Budget at a Glance workbook.

Why the source is vendored
--------------------------
`indiabudget.gov.in` publishes Budget at a Glance as a 28 KB XLSX alongside the
PDF. Committing that file (data/sources/) rather than fetching it at build time
means this extraction is reproducible in CI with no network, and means the exact
bytes the numbers came from are in the repo history. When next year's budget is
published, drop in the new workbook and re-run.

Why the numbers are not hand-typed
----------------------------------
Every figure on the explorer is a Union government number that a reader may act
on. Transcribing 60-odd values by hand is the kind of task that produces one
wrong digit nobody ever finds, so the page reads a generated file and `--check`
fails when that file drifts from the workbook. Same contract as
build-law-guide-cases.py.

What this does NOT establish
----------------------------
That the figures are correct. The Ministry's own disclaimer says the Excel "has
been generated from System. In case of any inconsistency, pdf files will be
treated as final." So this pipeline is faithful to the XLSX, and the XLSX is not
the authority. The explorer says so, and so does `meta.authoritative`.
"""
import argparse
import datetime
import re
import json
import pathlib
import sys

try:
    import openpyxl
except ImportError:
    print('FAIL - openpyxl is required: pip install openpyxl')
    sys.exit(1)

ROOT = pathlib.Path(__file__).resolve().parent.parent
BOOK = ROOT / 'data' / 'sources' / 'budget-at-a-glance-2026-27.xlsx'
OUT = ROOT / 'data' / 'union-budget.json'

# Column layout differs per sheet, so each block declares its own. Transfer of
# Resources genuinely publishes only THREE year columns -- it omits 2025-26
# Budget Estimates, which the other sheets carry. Forcing every block onto one
# four-year axis would invent a column that the Ministry did not publish, so
# each block keeps its own `years` and the page reads them per block.
#
# Rows are positional because the sheets have no stable keys: "Grand Total"
# appears twice on the expenditure sheet (closing the structural block, then the
# functional one), so matching on label alone silently picks the wrong row.
#
#   name: (sheet, first_row, last_row, col_english, col_first_year, n_years)
BLOCKS = {
    'receipts':              ('Receipts',              4, 34, 2, 3, 4),
    'expenditure_structure': ('Expenditure of GOI',    5, 14, 2, 3, 4),
    'expenditure_function':  ('Expenditure of GOI',   29, 53, 2, 3, 4),
    'deficit':               ('Deficit Statistics',    4, 38, 2, 3, 4),
    'transfers':             ('Transfer of Resources', 4, 35, 4, 5, 3),
}

# Lines that are group headers, not amounts -- "Subsidy -" sits above
# Fertiliser/Food/Petroleum and carries no value of its own. Kept in the output
# with children so the explorer can nest them, but never summed.
HEADERS = {'Subsidy -', "A. Centre's Expenditure", 'B. Transfers',
           'REVENUE RECEIPTS', '1. Tax Revenue', 'CAPITAL RECEIPTS'}


def num(v):
    """The workbook stores amounts as text. Return a float, or None if the cell
    is genuinely not a number (blank, or the '...' used for a nil charged
    provision) rather than coercing it to zero -- a nil and a not-applicable
    are different claims and the explorer renders them differently."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(',', '')
    if not s or s in ('...', '-', '—', '..'):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def years_of(ws, col_first, n_years):
    """Read year labels from the header.

    Two layouts appear in one workbook. Receipts and Transfer of Resources put
    year and kind in a single cell ("2025-26\\nबजट\\nअनुमान\\nBudget \\nEstimates").
    Expenditure of GOI and Deficit Statistics split them across two rows: bare
    years on one, "Actuals / Budget Estimates / Revised Estimates" on the next.
    Reading only the first layout labelled every expenditure column "Estimates",
    including the 2024-25 column that is Actuals -- which would have put a
    provisional label on a settled number.
    """
    rows = list(ws.iter_rows(min_row=1, max_row=9, values_only=True))
    for idx, row in enumerate(rows):
        cells = list(row) + [None] * 24
        cand = [cells[col_first + i] for i in range(n_years)]
        if not all(c and str(c).strip()[:2] == '20' for c in cand):
            continue

        def kinds_from(cs):
            out = []
            for c in cs:
                bits = [b.strip() for b in str(c or '').split('\n') if b.strip()]
                out.append(' '.join(' '.join(b.split()) for b in bits
                                    if re.fullmatch(r'[A-Za-z ]+', b)))
            return out

        kinds = kinds_from(cand)
        if not any(kinds) and idx + 1 < len(rows):
            nxt = list(rows[idx + 1]) + [None] * 24
            kinds = kinds_from([nxt[col_first + i] for i in range(n_years)])
        years = [str(c).split('\n')[0].strip() for c in cand]
        return [{'year': y, 'kind': k or 'Estimates'}
                for y, k in zip(years, kinds)]
    return []


def rows_of(ws, lo, hi, col_en, col_first, n_years):
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if not (lo <= i <= hi):
            continue
        cells = list(row) + [None] * 24
        label = cells[col_en]
        if label is None or not str(label).strip():
            continue
        raw = str(label).replace('\n', ' ')
        indent = len(raw) - len(raw.lstrip())   # leading spaces encode nesting
        name = ' '.join(raw.split())
        vals = [num(cells[col_first + k]) for k in range(n_years)]
        depth = 1 if indent >= 4 else 0
        if name in HEADERS or all(v is None for v in vals):
            out.append({'label': name, 'depth': depth, 'header': True})
        else:
            out.append({'label': name, 'depth': depth, 'values': vals})
    return out


def build():
    wb = openpyxl.load_workbook(BOOK, data_only=True, read_only=True)
    doc = {
        'meta': {
            'title': 'Union Budget of India — Budget at a Glance',
            'unit': '₹ crore',
            'source': {
                'name': 'Budget at a Glance, Union Budget 2026-27',
                'publisher': 'Ministry of Finance, Government of India',
                'url': 'https://www.indiabudget.gov.in/',
                'file': 'doc/Budget_at_Glance/budget_at_a_glance.xlsx',
            },
            # The Ministry's own words, quoted rather than paraphrased because a
            # paraphrase would soften it.
            'authoritative': (
                'The Ministry of Finance states: "The Budget data in excel has been '
                'generated from System. In case of any inconsistency, pdf files will '
                'be treated as final." These figures are transcribed from the Excel. '
                'For any figure you intend to rely on, check the PDF.'
            ),
            'extracted': datetime.date.today().isoformat(),
        },
        'blocks': {},
    }
    # The functional expenditure lines do not sum exactly to the published
    # Grand Total -- each line is rounded to the crore, so the residue runs
    # about +1 / -2 crore on a base of ~53 lakh crore. Computed here rather
    # than written on the page by hand, because a hand-typed reconciliation
    # note is precisely what went stale three times on the Law Docket. If a
    # future workbook drops a row, this number moves and the page says so.
    def reconcile(block_key):
        blk = doc['blocks'][block_key]
        rows = [r for r in blk['rows'] if 'values' in r]
        total = next((r for r in rows if r['label'] == 'Grand Total'), None)
        if not total:
            return None
        lines = [r for r in rows if r['label'] != 'Grand Total']
        out = []
        for i in range(len(blk['years'])):
            s = sum(r['values'][i] for r in lines if r['values'][i] is not None)
            out.append(round(s - total['values'][i], 2))
        return out

    for key, (sheet, lo, hi, col_en, col_first, n_years) in BLOCKS.items():
        ws = wb[sheet]
        doc['blocks'][key] = {
            'sheet': sheet,
            'years': years_of(ws, col_first, n_years),
            'rows': rows_of(ws, lo, hi, col_en, col_first, n_years),
        }
    doc['blocks']['expenditure_function']['residual'] = reconcile('expenditure_function')
    return doc


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--check', action='store_true',
                    help='fail if the committed JSON has drifted from the workbook')
    args = ap.parse_args()

    if not BOOK.exists():
        print('FAIL - %s not found.' % BOOK.relative_to(ROOT))
        return 1

    fresh = build()
    if args.check:
        if not OUT.exists():
            print('FAIL - %s not found. Run: python3 scripts/build-budget-data.py'
                  % OUT.relative_to(ROOT))
            return 1
        have = json.loads(OUT.read_text(encoding='utf-8'))
        # `extracted` is a build stamp, not data -- comparing it would make the
        # guard fail every day for no reason.
        a = dict(have); b = dict(fresh)
        a['meta'] = {k: v for k, v in a['meta'].items() if k != 'extracted'}
        b['meta'] = {k: v for k, v in b['meta'].items() if k != 'extracted'}
        if a != b:
            print('FAIL - data/union-budget.json is out of date with the workbook.')
            print('       Run: python3 scripts/build-budget-data.py')
            return 1
        n = sum(len(fresh['blocks'][k]['rows']) for k in BLOCKS)
        print('PASS - union-budget.json matches the workbook (%d lines across %d blocks).'
              % (n, len(BLOCKS)))
        return 0

    OUT.write_text(json.dumps(fresh, indent=2, ensure_ascii=False) + '\n',
                   encoding='utf-8')
    for k in BLOCKS:
        b = fresh['blocks'][k]
        vals = sum(1 for r in b['rows'] if 'values' in r)
        print('  %-24s %2d lines (%d with amounts, %d year columns)'
              % (k, len(b['rows']), vals, len(b['years'])))
    print('Wrote %s' % OUT.relative_to(ROOT))
    return 0


if __name__ == '__main__':
    sys.exit(main())
